from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from import_historical_market_data import METRICS, build_report, collect_rows, read_env


EXCLUDED_SHEETS = {
    "선물투자자별순매수금액": "investor_flow_deferred",
    "주식투자자별순매수금액": "investor_flow_deferred",
    "국공채형MMF": "mmf_deferred",
    "일반형MMF": "mmf_deferred",
}

PS_METRIC_PATTERN = re.compile(r"@\{\s*(?P<body>.*?)\s*\}", re.DOTALL)
PS_FIELD_PATTERN = re.compile(r"(?P<key>\w+)\s*=\s*(?P<value>\"[^\"]*\"|[0-9.]+)")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Check Excel source coverage for Daily Report mapped metrics.")
    parser.add_argument("--project-root", default=str(Path(__file__).resolve().parents[1]))
    parser.add_argument("--workbook")
    parser.add_argument("--report-date")
    parser.add_argument("--format", choices=["json", "markdown"], default="json")
    return parser.parse_args()


def resolve_workbook(project_root: Path, workbook: str | None) -> Path:
    if workbook:
        return Path(workbook).resolve()

    env = read_env(project_root)
    candidates = [
        env.get("INFOMAX_EXCEL_PATH"),
        project_root.parent.parent / "MARKET DAILY.xlsm",
        project_root.parent / "MARKET DAILY.xlsm",
        project_root / "MARKET DAILY.xlsm",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return Path(candidate).resolve()
    raise RuntimeError("MARKET DAILY.xlsm not found. Pass --workbook explicitly.")


def latest_report_date(rows_by_sheet: dict[str, dict[str, dict[str, Any]]]) -> str:
    core_keys = {"kospi", "usdkrw", "us_treasury_10y", "sp500", "wti"}
    metric_by_key = {metric.key: metric for metric in METRICS}
    counts: dict[str, int] = {}
    for key in core_keys:
        metric = metric_by_key[key]
        for row_date, row in rows_by_sheet.get(metric.sheet, {}).items():
            if row.get(key, {}).get("value") not in (None, 0):
                counts[row_date] = counts.get(row_date, 0) + 1
    candidates = [row_date for row_date, count in counts.items() if count >= 4]
    if not candidates:
        raise RuntimeError("No valid report date found from core metrics.")
    return sorted(candidates)[-1]


def workbook_sheet_summary(workbook_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True, keep_vba=True)
    summary: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        if sheet_name in {"MARKET DAILY", "camera"}:
            continue
        ws = wb[sheet_name]
        headers = []
        for column_index, cell in enumerate(ws[3], start=1):
            if column_index == 1:
                continue
            if cell.value not in (None, ""):
                headers.append(get_column_letter(column_index))
        summary.append(
            {
                "sheet": sheet_name,
                "rows": ws.max_row,
                "cols": ws.max_column,
                "data_columns": headers,
                "excluded_reason": EXCLUDED_SHEETS.get(sheet_name),
            }
        )
    return summary


def parse_powershell_metric_defs(project_root: Path) -> list[dict[str, Any]]:
    path = project_root / "scripts" / "Export-MarketDailyCachedValues.ps1"
    if not path.exists():
        return []

    text = path.read_text(encoding="utf-8-sig")
    defs: list[dict[str, Any]] = []
    for match in PS_METRIC_PATTERN.finditer(text):
        body = match.group("body")
        fields: dict[str, Any] = {}
        for field in PS_FIELD_PATTERN.finditer(body):
            value = field.group("value")
            if value.startswith('"') and value.endswith('"'):
                value = value[1:-1]
            fields[field.group("key")] = value
        if "Key" not in fields or "Sheet" not in fields or "Column" not in fields:
            continue
        defs.append(
            {
                "key": fields.get("Key"),
                "name": fields.get("Name"),
                "category": fields.get("Category"),
                "sheet": fields.get("Sheet"),
                "column": fields.get("Column"),
                "unit": fields.get("Unit"),
                "change_mode": fields.get("ChangeMode"),
                "value_multiplier": float(fields.get("ValueMultiplier", 1.0)),
            }
        )
    return defs


def mapping_parity(project_root: Path, metric_defs: list[dict[str, Any]]) -> dict[str, Any]:
    ps_defs = parse_powershell_metric_defs(project_root)
    python_by_key = {item["key"]: item for item in metric_defs}
    ps_by_key = {item["key"]: item for item in ps_defs}
    compared_fields = ["name", "category", "sheet", "column", "unit", "change_mode", "value_multiplier"]
    mismatches = []

    for key in sorted(set(python_by_key) & set(ps_by_key)):
        py_item = python_by_key[key]
        ps_item = ps_by_key[key]
        field_diffs = {
            field: {"python": py_item.get(field), "powershell": ps_item.get(field)}
            for field in compared_fields
            if py_item.get(field) != ps_item.get(field)
        }
        if field_diffs:
            mismatches.append({"key": key, "fields": field_diffs})

    return {
        "python_metric_count": len(metric_defs),
        "powershell_metric_count": len(ps_defs),
        "missing_in_powershell": sorted(set(python_by_key) - set(ps_by_key)),
        "missing_in_python": sorted(set(ps_by_key) - set(python_by_key)),
        "mismatches": mismatches,
    }


def main() -> int:
    args = parse_args()
    project_root = Path(args.project_root).resolve()
    workbook_path = resolve_workbook(project_root, args.workbook)
    rows_by_sheet = collect_rows(workbook_path)
    report_date = args.report_date or latest_report_date(rows_by_sheet)
    report = build_report(rows_by_sheet, report_date, workbook_path)
    observed_keys = {item["metric_key"] for item in report["observations"]}
    metric_defs = [asdict(metric) for metric in METRICS]
    mapped_keys = {metric["key"] for metric in metric_defs}
    parity = mapping_parity(project_root, metric_defs)

    mapped_by_sheet: dict[str, list[str]] = defaultdict(list)
    for metric in metric_defs:
        mapped_by_sheet[metric["sheet"]].append(f'{metric["column"]}:{metric["key"]}')

    sheet_summary = workbook_sheet_summary(workbook_path)
    mapped_sheets = {metric["sheet"] for metric in metric_defs}
    unmapped_source_sheets = [
        {
            "sheet": item["sheet"],
            "data_columns": item["data_columns"],
            "classification": item["excluded_reason"] or "unmapped_source_sheet",
        }
        for item in sheet_summary
        if item["sheet"] not in mapped_sheets and item["data_columns"]
    ]

    result = {
        "checked_at": datetime.now().isoformat(timespec="seconds"),
        "workbook": str(workbook_path),
        "report_date": report_date,
        "mapped_metric_count": len(metric_defs),
        "extracted_observation_count": len(report["observations"]),
        "missing_mapped_metrics": sorted(mapped_keys - observed_keys),
        "extra_observations": sorted(observed_keys - mapped_keys),
        "mapped_by_sheet": dict(sorted(mapped_by_sheet.items())),
        "unmapped_source_sheets": unmapped_source_sheets,
        "mapping_parity": parity,
    }

    if args.format == "json":
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print(f"# Excel Coverage Check ({report_date})")
        print()
        print(f"- Workbook: `{workbook_path}`")
        print(f"- Mapped metrics: {result['mapped_metric_count']}")
        print(f"- Extracted observations: {result['extracted_observation_count']}")
        print(f"- Missing mapped metrics: {len(result['missing_mapped_metrics'])}")
        print(f"- PowerShell mapping mismatches: {len(parity['mismatches'])}")
        print()
        print("## Unmapped Source Sheets")
        for item in unmapped_source_sheets:
            print(f"- `{item['sheet']}`: {item['classification']} ({len(item['data_columns'])} data columns)")
    has_parity_error = bool(parity["missing_in_powershell"] or parity["missing_in_python"] or parity["mismatches"])
    return 1 if result["missing_mapped_metrics"] or has_parity_error else 0


if __name__ == "__main__":
    raise SystemExit(main())
